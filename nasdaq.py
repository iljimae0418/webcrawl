#for collecting NASDAQ data 
import urllib 
from collections import defaultdict
import openpyxl 
from openpyxl import Workbook
from openpyxl import load_workbook
# get ticker names from file I/O 
f = open("nasdaqtickers.txt","r")
tickers = [] 
for line in f:
	newline = line.find("\n") 
	if newline == -1:  
		tickers.append(line)
	else:
		tickers.append(line[0:newline])
# web-crawling step 
data = defaultdict(list) 
for i in range(len(tickers)):
	name = tickers[i] #get the name of the ticker 
	link_front = "https://finance.google.com/finance/getprices?q="
	link_front = link_front+name 
	link_back = "&x=NASDAQ&i=1800&p=140d&f=d,c,o,h,l,v" 
	link = link_front + link_back 
	f = urllib.urlopen(link);  
	myfile = f.read() # this contains the contents of the link 
	idx = int(myfile.find("TIMEZONE_OFFSET=-240")+21) # note that TIMEZONE_OFFSET=-240 appears for all ticker urls. This formed the basis for my index calculations 
	numbers = myfile[idx:len(myfile)-1] # python substring 
	infoList = numbers.split(",") # this is convenient: separate by commas 
	#getting close[] data 
	close = []
	j = 1 
	while (j < len(infoList)): 
		close.append(infoList[j]) 
		j += 5 
	#getting volume[] data  
	volume = [] 
	k = 5
	while (k < len(infoList)): # code longer since we have to take care of newline (if you don't know what I'm talking about print infoList to check)
		process = infoList[k] 
		nlIdx = process.find("\n") #nlIdx = new line idx 
		if nlIdx == -1: 
			volume.append(process)
		else: 
			volume.append(process[0:nlIdx]) 
		k += 5 
	# if length of data is less than 200 skip  
	if len(close) < 200 and len(volume) < 200: 
		continue 
	#adding close[] data as the name of ticker 
	data.setdefault(name,[]).append(close) 
	#adding voluem[] data as the name + vol 
	data.setdefault(name+"vol",[]).append(volume) 
# work with excel file  
book = Workbook() 
sheet = book.active
# gives the second column data for a ticker k  
# v is going to be empty in the case there are less than 250 data 
# data format: (name of ticker, [list containing the 2nd column numbers])
# data format2: (name of ticker+vol, [list containing the last column numbers])
c = 1
for k,v in data.items():
	# flattening out list of list to list ... this logic is tricky 
 	temp = [item for sublist in v for item in sublist]
	# put ticker first  
	sheet.cell(row=1,column=c).value = k 
	r = 2 
	for i in range(len(temp)):
		sheet.cell(row=r,column=c).value = temp[i] 
		r += 1 
	c += 1 
# save the excel file 
book.save('nasdaqwebcrawl.xlsx')
